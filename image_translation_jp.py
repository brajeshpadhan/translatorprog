from PIL import Image, ImageDraw, ImageFont
import pytesseract
import deepl
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# ==============================
# CONFIGURATION
# ==============================
IMAGES_FOLDER = "images"
OUTPUT_FOLDER = "translated_images"
REPORT_XLSX = "japanese_text_report.xlsx"

DEEPL_API_KEY = "YOUR_DEEPL_API_KEY"
FONT_PATH = "arial.ttf"  # Path to TTF font
MAX_FONT_SIZE = 40        # Max font size for overlay

os.makedirs(OUTPUT_FOLDER, exist_ok=True)

translator = deepl.Translator(DEEPL_API_KEY)

# ==============================
# CREATE EXCEL REPORT
# ==============================
wb = Workbook()
ws = wb.active
ws.title = "OCR Translation Report"

# Header row
headers = ["Image File", "X", "Y", "Width", "Height", "Japanese Text", "English Translation"]
ws.append(headers)

# Add fill colors
japanese_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Yellow
english_fill = PatternFill(start_color="99CCFF", end_color="99CCFF", fill_type="solid")  # Blue

# ==============================
# STEP 1: Extract and Translate
# ==============================
report_rows = []

for filename in os.listdir(IMAGES_FOLDER):
    if filename.lower().endswith((".png", ".jpg", ".jpeg")):
        image_path = os.path.join(IMAGES_FOLDER, filename)
        img = Image.open(image_path)

        data = pytesseract.image_to_data(img, lang='jpn', output_type=pytesseract.Output.DICT)
        n_boxes = len(data['level'])

        for i in range(n_boxes):
            text = data['text'][i].strip()
            if text == "":
                continue

            x, y, w, h = data['left'][i], data['top'][i], data['width'][i], data['height'][i]

            # Translate Japanese to English
            try:
                translation = translator.translate_text(text, source_lang="JA", target_lang="EN-US").text
            except Exception as e:
                translation = text  # fallback
                print(f"Translation failed for '{text}': {e}")

            # Append to report
            row = [filename, x, y, w, h, text, translation]
            report_rows.append(row)

            ws.append(row)
            ws.cell(row=ws.max_row, column=6).fill = japanese_fill
            ws.cell(row=ws.max_row, column=7).fill = english_fill

# Save Excel report
wb.save(REPORT_XLSX)
print(f"Excel report generated: {REPORT_XLSX}")
print("Please review/validate the report before replacing text in images.\n")

# ==============================
# STEP 2: Replace text in images after validation
# ==============================
proceed = input("Do you want to replace Japanese text with English in images? (y/n): ").lower()
if proceed != "y":
    print("Process terminated. You can edit the report and rerun.")
    exit()

for row in report_rows:
    filename, x, y, w, h, japanese_text, english_text = row
    image_path = os.path.join(IMAGES_FOLDER, filename)
    output_path = os.path.join(OUTPUT_FOLDER, filename)

    img = Image.open(image_path)
    draw = ImageDraw.Draw(img)

    # Cover original Japanese text
    draw.rectangle([x, y, x + w, y + h], fill="white")

    # Adjust font size to fit bounding box width
    font_size = MAX_FONT_SIZE
    font = ImageFont.truetype(FONT_PATH, font_size)
    while font.getsize(english_text)[0] > w and font_size > 5:
        font_size -= 1
        font = ImageFont.truetype(FONT_PATH, font_size)

    # Draw English text
    draw.text((x, y), english_text, fill="black", font=font)

    img.save(output_path)
    print(f"Translated image saved: {output_path}")

print("\nAll images processed successfully!")

# pip install pillow pytesseract deepl openpyxl
#Tesseract OCR with Japanese language (jpn) installed:
#Ubuntu: sudo apt install tesseract-ocr tesseract-ocr-jpn
#Windows: Install Tesseract + Japanese traineddata
#MacOS: brew install tesseract-lang
